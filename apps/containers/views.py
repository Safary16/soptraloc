from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import AllowAny, IsAuthenticatedOrReadOnly
from django.utils import timezone
from django.db.models import Q
import tempfile
import os

from .models import Container
from .serializers import (
    ContainerSerializer, 
    ContainerListSerializer,
    ContainerStockExportSerializer
)
from .importers.embarque import EmbarqueImporter
from .importers.liberacion import LiberacionImporter
from .importers.programacion import ProgramacionImporter


class ContainerViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de contenedores
    """
    queryset = Container.objects.all()
    serializer_class = ContainerSerializer
    filterset_fields = ['estado', 'tipo', 'secuenciado', 'puerto', 'posicion_fisica']
    search_fields = ['container_id', 'nave', 'vendor', 'comuna']
    ordering_fields = ['created_at', 'fecha_programacion', 'fecha_liberacion']
    ordering = ['-created_at']
    permission_classes = [IsAuthenticatedOrReadOnly]
    
    def get_serializer_class(self):
        if self.action == 'list':
            return ContainerListSerializer
        return ContainerSerializer
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser], permission_classes=[AllowAny], url_path='import-embarque')
    def import_embarque(self, request):
        """
        Importa contenedores desde Excel de embarque
        Crea contenedores con estado 'por_arribar'
        
        NOTA: Este endpoint permite AllowAny por compatibilidad con sistemas externos.
        TODO: Cambiar a IsAuthenticated en producción para mayor seguridad.
        """
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No se proporcionó archivo'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        archivo = request.FILES['file']
        
        # Validar extensión de archivo
        if not archivo.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'Formato de archivo inválido. Solo se permiten archivos .xlsx o .xls'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar tamaño de archivo (máximo 10MB)
        max_size = 10 * 1024 * 1024  # 10MB en bytes
        if archivo.size > max_size:
            return Response(
                {'error': f'Archivo demasiado grande. Tamaño máximo: 10MB'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        usuario = request.user.username if request.user.is_authenticated else None
        
        # Guardar temporalmente
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            for chunk in archivo.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name
        
        try:
            # Procesar con el importador
            importer = EmbarqueImporter(tmp_path, usuario)
            resultados = importer.procesar()
            
            return Response({
                'success': True,
                'mensaje': f'Importación completada',
                'creados': resultados['creados'],
                'actualizados': resultados['actualizados'],
                'errores': resultados['errores'],
                'detalles': resultados['detalles']
            })
        
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error importando embarque: {str(e)}", exc_info=True)
            return Response(
                {'error': 'Error procesando el archivo. Verifique el formato y vuelva a intentar.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        finally:
            # Limpiar archivo temporal
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser], permission_classes=[AllowAny], url_path='import-liberacion')
    def import_liberacion(self, request):
        """
        Importa liberaciones desde Excel
        Actualiza contenedores a estado 'liberado' y asigna posición física
        
        NOTA: Este endpoint permite AllowAny por compatibilidad con sistemas externos.
        TODO: Cambiar a IsAuthenticated en producción para mayor seguridad.
        """
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No se proporcionó archivo'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        archivo = request.FILES['file']
        
        # Validar extensión de archivo
        if not archivo.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'Formato de archivo inválido. Solo se permiten archivos .xlsx o .xls'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar tamaño de archivo (máximo 10MB)
        max_size = 10 * 1024 * 1024  # 10MB en bytes
        if archivo.size > max_size:
            return Response(
                {'error': f'Archivo demasiado grande. Tamaño máximo: 10MB'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        usuario = request.user.username if request.user.is_authenticated else None
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            for chunk in archivo.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name
        
        try:
            importer = LiberacionImporter(tmp_path, usuario)
            resultados = importer.procesar()
            
            return Response({
                'success': True,
                'mensaje': f'Importación de liberación completada',
                'liberados': resultados['liberados'],
                'por_liberar': resultados.get('por_liberar', 0),
                'no_encontrados': resultados['no_encontrados'],
                'errores': resultados['errores'],
                'detalles': resultados['detalles']
            })
        
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser], permission_classes=[AllowAny], url_path='import-programacion')
    def import_programacion(self, request):
        """
        Importa programaciones desde Excel
        Crea programaciones y actualiza contenedores a 'programado'
        
        NOTA: Este endpoint permite AllowAny por compatibilidad con sistemas externos.
        TODO: Cambiar a IsAuthenticated en producción para mayor seguridad.
        """
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No se proporcionó archivo'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        archivo = request.FILES['file']
        
        # Validar extensión de archivo
        if not archivo.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'Formato de archivo inválido. Solo se permiten archivos .xlsx o .xls'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar tamaño de archivo (máximo 10MB)
        max_size = 10 * 1024 * 1024  # 10MB en bytes
        if archivo.size > max_size:
            return Response(
                {'error': f'Archivo demasiado grande. Tamaño máximo: 10MB'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        usuario = request.user.username if request.user.is_authenticated else None
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            for chunk in archivo.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name
        
        try:
            importer = ProgramacionImporter(tmp_path, usuario)
            resultados = importer.procesar()
            
            return Response({
                'success': True,
                'mensaje': f'Importación de programación completada',
                'programados': resultados['programados'],
                'no_encontrados': resultados['no_encontrados'],
                'cd_no_encontrado': resultados['cd_no_encontrado'],
                'errores': resultados['errores'],
                'alertas_generadas': resultados['alertas_generadas'],
                'detalles': resultados['detalles']
            })
        
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    @action(detail=False, methods=['get'], url_path='export-stock')
    def export_stock(self, request):
        """
        Exporta stock de contenedores liberados y por arribar (formato JSON)
        Incluye flag de 'secuenciado' para próximas liberaciones
        """
        # Filtrar solo liberados y por_arribar
        containers = Container.objects.filter(
            Q(estado='liberado') | Q(estado='por_arribar')
        ).order_by('secuenciado', '-fecha_liberacion')
        
        serializer = ContainerStockExportSerializer(containers, many=True)
        
        return Response({
            'success': True,
            'total': containers.count(),
            'containers': serializer.data
        })
    
    @action(detail=False, methods=['get'], url_path='export-liberacion-excel')
    def export_liberacion_excel(self, request):
        """
        Exporta contenedores liberados y por liberar a Excel
        Incluye: ID, Nave, Estado, Peso Total, Contenido, Demurrage, etc.
        """
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from datetime import datetime
        
        # Filtrar contenedores liberados y por arribar
        containers = Container.objects.filter(
            Q(estado='liberado') | Q(estado='por_arribar')
        ).select_related('cd_entrega').order_by('-fecha_demurrage', 'estado')
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Liberados y Por Liberar"
        
        # Estilos
        header_fill = PatternFill(start_color="E95420", end_color="E95420", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            'CONTAINER ID', 'ESTADO', 'NAVE', 'TIPO', 'TIPO CARGA',
            'PESO CARGA (KG)', 'TARA (KG)', 'PESO TOTAL (KG)', 'PESO TOTAL (TON)',
            'CONTENIDO', 'POSICIÓN FÍSICA', 'PUERTO',
            'FECHA DEMURRAGE', 'DÍAS DEMURRAGE', 'URGENCIA',
            'CD ENTREGA', 'COMUNA', 'VENDOR', 'SELLO',
            'FECHA LIBERACIÓN', 'FECHA ETA', 'SECUENCIADO'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Datos
        row = 2
        for container in containers:
            ws.cell(row=row, column=1, value=container.container_id).border = border
            ws.cell(row=row, column=2, value=container.get_estado_display()).border = border
            ws.cell(row=row, column=3, value=container.nave).border = border
            ws.cell(row=row, column=4, value=container.get_tipo_display()).border = border
            ws.cell(row=row, column=5, value=container.get_tipo_carga_display()).border = border
            
            ws.cell(row=row, column=6, value=float(container.peso_carga or 0)).border = border
            ws.cell(row=row, column=7, value=float(container.tara or 0)).border = border
            ws.cell(row=row, column=8, value=container.peso_total).border = border
            ws.cell(row=row, column=9, value=round(container.peso_total / 1000, 2)).border = border
            
            ws.cell(row=row, column=10, value=container.contenido or '').border = border
            ws.cell(row=row, column=11, value=container.posicion_fisica or '').border = border
            ws.cell(row=row, column=12, value=container.puerto).border = border
            
            # Demurrage
            if container.fecha_demurrage:
                ws.cell(row=row, column=13, value=container.fecha_demurrage.strftime('%d/%m/%Y')).border = border
                dias = container.dias_para_demurrage
                ws.cell(row=row, column=14, value=dias if dias is not None else '').border = border
                ws.cell(row=row, column=15, value=container.urgencia_demurrage.upper()).border = border
                
                # Colorear según urgencia
                urgencia_cell = ws.cell(row=row, column=15)
                if container.urgencia_demurrage == 'vencido':
                    urgencia_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    urgencia_cell.font = Font(bold=True, color="FFFFFF")
                elif container.urgencia_demurrage == 'critico':
                    urgencia_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                    urgencia_cell.font = Font(bold=True, color="FFFFFF")
                elif container.urgencia_demurrage == 'alto':
                    urgencia_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                elif container.urgencia_demurrage == 'medio':
                    urgencia_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            else:
                ws.cell(row=row, column=13, value='').border = border
                ws.cell(row=row, column=14, value='').border = border
                ws.cell(row=row, column=15, value='SIN FECHA').border = border
            
            ws.cell(row=row, column=16, value=container.cd_entrega.nombre if container.cd_entrega else '').border = border
            ws.cell(row=row, column=17, value=container.comuna or '').border = border
            ws.cell(row=row, column=18, value=container.vendor or '').border = border
            ws.cell(row=row, column=19, value=container.sello or '').border = border
            
            # Fechas
            if container.fecha_liberacion:
                ws.cell(row=row, column=20, value=container.fecha_liberacion.strftime('%d/%m/%Y %H:%M')).border = border
            else:
                ws.cell(row=row, column=20, value='').border = border
            
            if container.fecha_eta:
                ws.cell(row=row, column=21, value=container.fecha_eta.strftime('%d/%m/%Y')).border = border
            else:
                ws.cell(row=row, column=21, value='').border = border
            
            ws.cell(row=row, column=22, value='SÍ' if container.secuenciado else 'NO').border = border
            
            row += 1
        
        # Ajustar anchos de columna
        column_widths = {
            1: 18, 2: 15, 3: 25, 4: 10, 5: 15,
            6: 15, 7: 12, 8: 15, 9: 15, 10: 40,
            11: 18, 12: 15, 13: 18, 14: 15, 15: 12,
            16: 30, 17: 15, 18: 30, 19: 15, 20: 18,
            21: 15, 22: 12
        }
        for col, width in column_widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'liberados_por_liberar_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    @action(detail=True, methods=['post'])
    def cambiar_estado(self, request, pk=None):
        """
        Cambia manualmente el estado de un contenedor
        """
        container = self.get_object()
        nuevo_estado = request.data.get('estado')
        
        if not nuevo_estado:
            return Response(
                {'error': 'Estado no proporcionado'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if nuevo_estado not in dict(Container.ESTADOS):
            return Response(
                {'error': f'Estado inválido: {nuevo_estado}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        usuario = request.user.username if request.user.is_authenticated else None
        container.cambiar_estado(nuevo_estado, usuario)
        
        serializer = self.get_serializer(container)
        return Response({
            'success': True,
            'mensaje': f'Estado cambiado a {container.get_estado_display()}',
            'container': serializer.data
        })
    
    @action(detail=True, methods=['post'])
    def marcar_liberado(self, request, pk=None):
        """Marca contenedor como liberado por aduana/naviera"""
        container = self.get_object()
        usuario = request.user.username if request.user.is_authenticated else None
        container.cambiar_estado('liberado', usuario)
        
        serializer = self.get_serializer(container)
        return Response({
            'success': True,
            'mensaje': f'Contenedor {container.container_id} liberado',
            'container': serializer.data
        })
    
    @action(detail=True, methods=['post'])
    def programar(self, request, pk=None):
        """
        Programa un contenedor manualmente desde operaciones
        Crea la programación y actualiza el estado del contenedor
        
        Payload:
        {
            "cd_id": 1,
            "fecha_programada": "2025-11-10T14:00:00Z",
            "cliente": "Cliente XYZ",
            "observaciones": "..."
        }
        """
        from apps.programaciones.models import Programacion
        from apps.cds.models import CD
        from apps.events.models import Event
        
        container = self.get_object()
        
        # Validar que esté liberado
        if container.estado not in ['liberado', 'secuenciado']:
            return Response(
                {'error': f'Contenedor debe estar liberado o secuenciado. Estado actual: {container.get_estado_display()}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar datos requeridos
        cd_id = request.data.get('cd_id')
        fecha_programada = request.data.get('fecha_programada')
        
        if not cd_id:
            return Response(
                {'error': 'cd_id es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not fecha_programada:
            return Response(
                {'error': 'fecha_programada es requerida'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Obtener CD
        try:
            cd = CD.objects.get(id=cd_id, activo=True)
        except CD.DoesNotExist:
            return Response(
                {'error': f'CD con ID {cd_id} no encontrado o inactivo'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Parsear fecha
        from dateutil import parser as date_parser
        try:
            fecha_programada_dt = date_parser.parse(fecha_programada)
        except Exception:
            return Response(
                {'error': 'Formato de fecha inválido. Use formato ISO 8601: YYYY-MM-DDTHH:MM:SSZ'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Verificar si ya existe una programación
        if hasattr(container, 'programacion') and container.programacion:
            return Response(
                {'error': 'Este contenedor ya tiene una programación asociada'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Crear programación
        programacion = Programacion.objects.create(
            container=container,
            cd=cd,
            fecha_programada=fecha_programada_dt,
            cliente=request.data.get('cliente', container.cliente or ''),
            direccion_entrega=cd.direccion,
            observaciones=request.data.get('observaciones', '')
        )
        
        # Actualizar contenedor
        container.cd_entrega = cd
        container.fecha_programacion = timezone.now()
        container.cambiar_estado('programado', request.user.username if request.user.is_authenticated else None)
        
        # Crear evento de auditoría
        Event.objects.create(
            container=container,
            event_type='programacion_manual',
            detalles={
                'cd': cd.nombre,
                'fecha_programada': fecha_programada_dt.isoformat(),
                'cliente': programacion.cliente,
                'programacion_id': programacion.id
            },
            usuario=request.user.username if request.user.is_authenticated else None
        )
        
        serializer = self.get_serializer(container)
        return Response({
            'success': True,
            'mensaje': f'Contenedor {container.container_id} programado para {cd.nombre}',
            'container': serializer.data,
            'programacion_id': programacion.id,
            'fecha_programada': fecha_programada_dt.isoformat()
        }, status=status.HTTP_201_CREATED)
    
    @action(detail=False, methods=['get'], url_path='liberados')
    def liberados(self, request):
        """
        Lista contenedores liberados disponibles para programar
        """
        containers = Container.objects.filter(
            estado='liberado'
        ).select_related('cd_entrega').order_by('-fecha_liberacion')
        
        serializer = ContainerListSerializer(containers, many=True)
        
        return Response({
            'success': True,
            'total': containers.count(),
            'containers': serializer.data
        })
    
    @action(detail=True, methods=['post'])
    def marcar_vacio(self, request, pk=None):
        """Marca contenedor como vacío (descargado, esperando retiro)"""
        container = self.get_object()
        
        if container.estado != 'descargado':
            return Response(
                {'error': f'Contenedor debe estar descargado. Estado actual: {container.get_estado_display()}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        usuario = request.user.username if request.user.is_authenticated else None
        container.cambiar_estado('vacio', usuario)
        
        serializer = self.get_serializer(container)
        return Response({
            'success': True,
            'mensaje': f'Contenedor {container.container_id} marcado como vacío',
            'container': serializer.data
        })
    
    @action(detail=True, methods=['post'])
    def iniciar_retorno(self, request, pk=None):
        """Inicia retorno de contenedor vacío a depósito"""
        container = self.get_object()
        
        if container.estado != 'vacio':
            return Response(
                {'error': f'Contenedor debe estar vacío. Estado actual: {container.get_estado_display()}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        usuario = request.user.username if request.user.is_authenticated else None
        container.cambiar_estado('vacio_en_ruta', usuario)
        
        serializer = self.get_serializer(container)
        return Response({
            'success': True,
            'mensaje': f'Retorno iniciado para {container.container_id}',
            'container': serializer.data
        })
    
    @action(detail=True, methods=['post'])
    def marcar_devuelto(self, request, pk=None):
        """Marca contenedor como devuelto a depósito naviera"""
        container = self.get_object()
        
        if container.estado != 'vacio_en_ruta':
            return Response(
                {'error': f'Contenedor debe estar en ruta vacío. Estado actual: {container.get_estado_display()}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        usuario = request.user.username if request.user.is_authenticated else None
        container.cambiar_estado('devuelto', usuario)
        
        serializer = self.get_serializer(container)
        return Response({
            'success': True,
            'mensaje': f'Contenedor {container.container_id} devuelto a depósito',
            'container': serializer.data
        })
    
    @action(detail=True, methods=['post'])
    def registrar_arribo(self, request, pk=None):
        """
        Registra el arribo manual de un contenedor al CD
        Cambia estado a 'entregado'
        """
        container = self.get_object()
        
        if container.estado != 'en_ruta':
            return Response(
                {'error': f'Contenedor debe estar en_ruta. Estado actual: {container.get_estado_display()}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        usuario = request.user.username if request.user.is_authenticated else None
        container.cambiar_estado('entregado', usuario)
        
        serializer = self.get_serializer(container)
        return Response({
            'success': True,
            'mensaje': f'Arribo registrado para {container.container_id}',
            'container': serializer.data
        })
    
    @action(detail=True, methods=['post'])
    def registrar_descarga(self, request, pk=None):
        """
        Registra la descarga del contenedor en el CD
        Cambia estado a 'descargado'
        
        Lógica de negocio:
        - Si cd.requiere_espera_carga=True (Puerto Madero, Campos, Quilicura):
          conductor espera descarga sobre camión, luego retorna a CCTI/depot con vacío
        - Si cd.permite_soltar_contenedor=True (El Peñón):
          drop & hook, conductor libre inmediatamente, puede recoger otro vacío
        
        Automáticamente crea registro TiempoOperacion para ML
        """
        from django.utils import timezone
        from apps.cds.models import CD
        from apps.programaciones.models import TiempoOperacion
        from datetime import timedelta
        
        container = self.get_object()
        
        if container.estado != 'entregado':
            return Response(
                {'error': f'Contenedor debe estar entregado. Estado actual: {container.get_estado_display()}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Registrar hora de descarga
        hora_fin = timezone.now()
        container.hora_descarga = hora_fin
        usuario = request.user.username if request.user.is_authenticated else None
        container.cambiar_estado('descargado', usuario)
        
        # Verificar configuración del CD
        cd = container.cd_entrega
        mensaje_adicional = ""
        
        if cd:
            if cd.requiere_espera_carga:
                mensaje_adicional = f" (Conductor debe esperar descarga sobre camión en {cd.nombre})"
            elif cd.permite_soltar_contenedor:
                mensaje_adicional = f" (Drop & hook en {cd.nombre}, conductor libre inmediatamente)"
            
            # 🆕 Crear registro ML de TiempoOperacion
            try:
                # Obtener programación y conductor
                programacion = getattr(container, 'programacion', None)
                conductor = programacion.driver if programacion else None
                
                # Calcular hora de inicio (usar fecha_entrega como aproximación)
                hora_inicio = container.fecha_entrega if container.fecha_entrega else (hora_fin - timedelta(hours=1))
                
                # Calcular tiempos
                tiempo_real_min = int((hora_fin - hora_inicio).total_seconds() / 60)
                tiempo_estimado = cd.tiempo_promedio_descarga_min or 60
                
                # Crear registro
                TiempoOperacion.objects.create(
                    cd=cd,
                    conductor=conductor,
                    container=container,
                    tipo_operacion='descarga_cd',
                    tiempo_estimado_min=tiempo_estimado,
                    tiempo_real_min=tiempo_real_min,
                    hora_inicio=hora_inicio,
                    hora_fin=hora_fin,
                    observaciones=f"Auto-generado desde registrar_descarga por {usuario or 'sistema'}"
                )
                
                mensaje_adicional += f" [ML: {tiempo_real_min}min registrados]"
            
            except Exception as e:
                # No fallar si el registro ML falla
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f"Error creando registro TiempoOperacion: {str(e)}")
        
        serializer = self.get_serializer(container)
        return Response({
            'success': True,
            'mensaje': f'Descarga registrada para {container.container_id}{mensaje_adicional}',
            'container': serializer.data,
            'hora_descarga': container.hora_descarga.isoformat() if container.hora_descarga else None
        })
    
    @action(detail=True, methods=['post'])
    def soltar_contenedor(self, request, pk=None):
        """
        Permite soltar un contenedor en El Peñón (drop & hook)
        Solo funciona si cd.permite_soltar_contenedor=True
        Cambia estado a 'descargado' y libera al conductor inmediatamente
        """
        from django.utils import timezone
        
        container = self.get_object()
        
        if container.estado != 'entregado':
            return Response(
                {'error': f'Contenedor debe estar entregado. Estado actual: {container.get_estado_display()}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        cd = container.cd_entrega
        if not cd:
            return Response(
                {'error': 'Contenedor no tiene CD de entrega asignado'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not cd.permite_soltar_contenedor:
            return Response(
                {'error': f'CD {cd.nombre} no permite drop & hook. Solo El Peñón tiene esta opción.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Soltar contenedor
        container.hora_descarga = timezone.now()
        usuario = request.user.username if request.user.is_authenticated else None
        container.cambiar_estado('descargado', usuario)
        
        # El CD recibe el vacío automáticamente (esto se maneja con signals más tarde)
        # Por ahora solo registramos el evento
        
        serializer = self.get_serializer(container)
        return Response({
            'success': True,
            'mensaje': f'Contenedor soltado en {cd.nombre}. Conductor liberado inmediatamente.',
            'container': serializer.data,
            'conductor_libre': True
        })
